"""Gera a Calculadora de Preços esDEV (.xlsx) com fórmulas nativas de Excel.

Uso:
    python scripts/gerar_calculadora.py [ficheiro_de_saida.xlsx]

O ficheiro gerado tem 4 folhas:
    Calculadora  -> inputs (amarelo) + resultados
    Detalhe      -> horas e valor por fase (auditoria do cálculo)
    Parametros   -> todas as tarifas, fatores e tabelas (é aqui que se afina o modelo)
    Proposta     -> resumo pronto a copiar para a proposta ao cliente

Nada é hardcoded como valor: os resultados são fórmulas, logo o Excel recalcula
sozinho quando se mexe nos inputs ou nos parâmetros.
"""

from __future__ import annotations

import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

# --- Identidade visual -------------------------------------------------------
AZUL = "0F2E4C"
AZUL_CLARO = "E8F0F7"
CINZA = "F2F2F2"
AMARELO = "FFF3C4"
VERDE = "E4F3E8"
BRANCO = "FFFFFF"

EUR = '#,##0 "€"'
EUR2 = '#,##0.00 "€"'
PCT = "0%"
HORAS = '#,##0.0 "h"'
FATOR = "0.00"

fill_header = PatternFill("solid", fgColor=AZUL)
fill_sub = PatternFill("solid", fgColor=AZUL_CLARO)
fill_input = PatternFill("solid", fgColor=AMARELO)
fill_out = PatternFill("solid", fgColor=VERDE)
fill_zebra = PatternFill("solid", fgColor=CINZA)

f_title = Font(bold=True, size=16, color=AZUL)
f_header = Font(bold=True, size=11, color=BRANCO)
f_sub = Font(bold=True, size=11, color=AZUL)
f_label = Font(size=11)
f_bold = Font(bold=True, size=11)
f_big = Font(bold=True, size=14, color=AZUL)
f_hint = Font(size=9, italic=True, color="777777")

thin = Side(style="thin", color="BFBFBF")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

# --- Dados do modelo ---------------------------------------------------------
TARIFAS = [
    ("Descoberta / Estratégia", 40),
    ("UX / Design", 45),
    ("Frontend", 45),
    ("Backend / CMS", 50),
    ("Integrações / API", 55),
    ("QA / Testes", 35),
    ("Gestão de projeto", 40),
    ("Produção de conteúdos", 40),
    ("SEO", 45),
]

CONSTANTES = [
    ("Gestão de projeto (% das horas técnicas)", 0.15, PCT),
    ("Buffer de risco (% sobre horas e valor)", 0.10, PCT),
    ("Capacidade de produção (h / semana)", 25, "0"),
    ("Fator do Preço Mínimo", 0.85, FATOR),
    ("Fator do Preço Premium", 1.30, FATOR),
    ("Horas por página / ecrã extra", 3, "0.0"),
    ("Acréscimo por idioma adicional (%)", 0.12, PCT),
    ("Horas por integração externa extra", 6, "0.0"),
    ("Arredondamento do preço final (€)", 25, "0"),
    ("IVA", 0.23, PCT),
]

# Tipo, Descoberta, UX/Design, Frontend, Backend, Integrações, QA, Páginas incl., Preço mín. (€)
TIPOS = [
    ("Landing page (1 página)", 2, 6, 14, 0, 2, 3, 1, 450),
    ("Website institucional (até 5 páginas)", 4, 12, 28, 4, 4, 6, 5, 950),
    ("Website + Blog / CMS", 6, 16, 40, 16, 6, 10, 8, 1500),
    ("Loja online (e-commerce)", 10, 24, 70, 40, 20, 18, 10, 2900),
    ("Web app / Dashboard", 14, 24, 90, 90, 24, 28, 8, 4500),
    ("Automação / Integrações", 8, 4, 10, 40, 40, 12, 1, 1500),
    ("Redesign / Recuperação de site", 4, 10, 24, 10, 6, 8, 5, 900),
]

COMPLEXIDADE = [
    ("Simples — pouco conteúdo, sem lógica", 0.85),
    ("Média — padrão de mercado", 1.00),
    ("Alta — regras de negócio próprias", 1.35),
    ("Muito alta — crítico / à medida", 1.70),
]

DESIGN = [
    ("Template adaptado", 0.70),
    ("Semi-custom (base + identidade)", 1.00),
    ("100% à medida", 1.40),
]

URGENCIA = [
    ("Normal (prazo confortável)", 1.00),
    ("Apertado (-25% de prazo)", 1.20),
    ("Urgente (fora de horas / fila)", 1.40),
]

CLIENTE = [
    ("Startup / Negócio local", 0.90),
    ("PME", 1.00),
    ("Corporate / Institucional", 1.20),
]

CONTEUDOS = [
    ("Cliente fornece textos e imagens", 0),
    ("esDEV produz textos e imagens", 10),
]

SEO = [
    ("Sem SEO", 0),
    ("SEO base (on-page + meta)", 6),
    ("SEO avançado (estrutura + performance)", 16),
]

# Plano, % do preço recomendado / mês, mínimo €/mês, inclui
MANUTENCAO = [
    (
        "Essencial",
        0.020,
        39,
        "Alojamento, updates de segurança, backups, 1h/mês de alterações, resposta em 3 dias úteis.",
    ),
    (
        "Pro",
        0.035,
        89,
        "Tudo do Essencial + monitorização, 3h/mês de alterações, relatório trimestral, resposta em 1 dia útil.",
    ),
    (
        "Premium",
        0.055,
        149,
        "Tudo do Pro + 6h/mês de evolução, prioridade absoluta, SLA 4h úteis, reunião mensal.",
    ),
]

DEFAULTS = {
    "cliente": "Cliente / Nome do projeto",
    "tipo": TIPOS[1][0],
    "complexidade": COMPLEXIDADE[1][0],
    "paginas": 5,
    "design": DESIGN[1][0],
    "idiomas": 1,
    "integracoes": 0,
    "conteudos": CONTEUDOS[0][0],
    "seo": SEO[1][0],
    "urgencia": URGENCIA[0][0],
    "perfil": CLIENTE[1][0],
    "desconto": 0.0,
    "custos": 0,
}


def secao(ws, row, texto, ncols=4, col=1):
    """Faixa de secção a toda a largura."""
    ws.cell(row=row, column=col, value=texto).font = f_header
    for c in range(col, col + ncols):
        ws.cell(row=row, column=c).fill = fill_header
    ws.row_dimensions[row].height = 20


def rotulo(ws, row, texto, dica=None, col=2):
    ws.cell(row=row, column=col, value=texto).font = f_label
    if dica:
        ws.cell(row=row, column=col + 2, value=dica).font = f_hint


def build_parametros(ws):
    ws.sheet_view.showGridLines = False
    ws["A1"] = "PARÂMETROS DO MODELO esDEV"
    ws["A1"].font = f_title
    ws["A2"] = (
        "Toda a calibração do preço vive nesta folha. Altera aqui — nunca nas fórmulas da folha Calculadora."
    )
    ws["A2"].font = f_hint

    secao(ws, 3, "TARIFAS POR FASE (€ / hora)", ncols=3)
    for i, (nome, valor) in enumerate(TARIFAS):
        r = 4 + i
        ws.cell(row=r, column=1, value=nome).font = f_label
        c = ws.cell(row=r, column=2, value=valor)
        c.font = f_bold
        c.fill = fill_input
        c.number_format = EUR
        c.border = box

    secao(ws, 14, "CONSTANTES DE CÁLCULO", ncols=3)
    for i, (nome, valor, fmt) in enumerate(CONSTANTES):
        r = 15 + i
        ws.cell(row=r, column=1, value=nome).font = f_label
        c = ws.cell(row=r, column=2, value=valor)
        c.font = f_bold
        c.fill = fill_input
        c.number_format = fmt
        c.border = box

    secao(ws, 26, "TIPOS DE PROJETO — horas base por fase", ncols=9)
    cabecalho = [
        "Tipo de projeto",
        "Descoberta",
        "UX / Design",
        "Frontend",
        "Backend",
        "Integrações",
        "QA",
        "Páginas incl.",
        "Preço mínimo",
    ]
    for j, titulo in enumerate(cabecalho):
        c = ws.cell(row=26, column=1 + j, value=titulo)
        c.font = f_header
        c.fill = fill_header
        c.alignment = Alignment(wrap_text=True, vertical="center")
    for i, linha in enumerate(TIPOS):
        r = 27 + i
        for j, valor in enumerate(linha):
            c = ws.cell(row=r, column=1 + j, value=valor)
            c.border = box
            if j == 0:
                c.font = f_label
            else:
                c.fill = fill_input
                c.number_format = EUR if j == 8 else "0.0"

    def tabela_fator(row_titulo, titulo, dados, fmt=FATOR, col_b="Fator"):
        secao(ws, row_titulo, titulo, ncols=2)
        ws.cell(row=row_titulo, column=2, value=col_b).font = f_header
        for i, (nome, valor) in enumerate(dados):
            r = row_titulo + 1 + i
            ws.cell(row=r, column=1, value=nome).font = f_label
            c = ws.cell(row=r, column=2, value=valor)
            c.fill = fill_input
            c.number_format = fmt
            c.border = box

    tabela_fator(35, "NÍVEL DE COMPLEXIDADE", COMPLEXIDADE)
    tabela_fator(41, "NÍVEL DE DESIGN", DESIGN)
    tabela_fator(46, "PRAZO / URGÊNCIA", URGENCIA)
    tabela_fator(51, "PERFIL DE CLIENTE", CLIENTE)
    tabela_fator(56, "PRODUÇÃO DE CONTEÚDOS", CONTEUDOS, fmt="0.0", col_b="Horas")
    tabela_fator(60, "SEO", SEO, fmt="0.0", col_b="Horas")

    secao(ws, 65, "PLANOS DE MANUTENÇÃO MENSAL", ncols=4)
    for j, titulo in enumerate(["Plano", "% do preço recomendado", "Mínimo / mês", "Inclui"]):
        c = ws.cell(row=65, column=1 + j, value=titulo)
        c.font = f_header
        c.fill = fill_header
    for i, (nome, pct, minimo, inclui) in enumerate(MANUTENCAO):
        r = 66 + i
        ws.cell(row=r, column=1, value=nome).font = f_bold
        c = ws.cell(row=r, column=2, value=pct)
        c.number_format = "0.0%"
        c.fill = fill_input
        c.border = box
        c = ws.cell(row=r, column=3, value=minimo)
        c.number_format = EUR
        c.fill = fill_input
        c.border = box
        ws.cell(row=r, column=4, value=inclui).font = f_label

    larguras = {"A": 42, "B": 16, "C": 14, "D": 14, "E": 14, "F": 14, "G": 12, "H": 13, "I": 14}
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w


# Referências reutilizadas nas fórmulas
P = "Parametros!"
T_TIPOS = f"{P}$A$27:$I$33"
T_COMPLEX = f"{P}$A$36:$B$39"
T_DESIGN = f"{P}$A$42:$B$44"
T_URG = f"{P}$A$47:$B$49"
T_CLIENTE = f"{P}$A$52:$B$54"
T_CONT = f"{P}$A$57:$B$58"
T_SEO = f"{P}$A$61:$B$63"

IN_TIPO = "Calculadora!$C$6"
IN_COMPLEX = "Calculadora!$C$7"
IN_PAGINAS = "Calculadora!$C$8"
IN_DESIGN = "Calculadora!$C$9"
IN_IDIOMAS = "Calculadora!$C$10"
IN_INTEGR = "Calculadora!$C$11"
IN_CONT = "Calculadora!$C$12"
IN_SEO = "Calculadora!$C$13"
IN_URG = "Calculadora!$C$14"
IN_PERFIL = "Calculadora!$C$15"

F_COMPLEX = f"VLOOKUP({IN_COMPLEX},{T_COMPLEX},2,FALSE)"
F_DESIGN = f"VLOOKUP({IN_DESIGN},{T_DESIGN},2,FALSE)"
F_IDIOMAS = f"(1+MAX(0,{IN_IDIOMAS}-1)*{P}$B$21)"
PAG_EXTRA = f"MAX(0,{IN_PAGINAS}-VLOOKUP({IN_TIPO},{T_TIPOS},8,FALSE))*{P}$B$20"


def build_detalhe(ws):
    ws.sheet_view.showGridLines = False
    ws["A1"] = "DETALHE DO CÁLCULO"
    ws["A1"].font = f_title
    ws["A2"] = "Folha de auditoria: mostra de onde vem cada hora e cada euro. Não requer edição."
    ws["A2"].font = f_hint

    for j, titulo in enumerate(
        ["Fase", "Horas base", "Fator", "Horas ajustadas", "Tarifa €/h", "Valor"]
    ):
        c = ws.cell(row=4, column=1 + j, value=titulo)
        c.font = f_header
        c.fill = fill_header
        c.alignment = Alignment(wrap_text=True, vertical="center")

    fases = [
        (
            "Descoberta / Estratégia",
            f"=VLOOKUP({IN_TIPO},{T_TIPOS},2,FALSE)",
            f"={F_COMPLEX}",
            f"={P}$B$4",
        ),
        (
            "UX / Design",
            f"=VLOOKUP({IN_TIPO},{T_TIPOS},3,FALSE)+{PAG_EXTRA}*0.4",
            f"={F_COMPLEX}*{F_DESIGN}",
            f"={P}$B$5",
        ),
        (
            "Frontend",
            f"=VLOOKUP({IN_TIPO},{T_TIPOS},4,FALSE)+{PAG_EXTRA}*0.6",
            f"={F_COMPLEX}*{F_DESIGN}*{F_IDIOMAS}",
            f"={P}$B$6",
        ),
        (
            "Backend / CMS",
            f"=VLOOKUP({IN_TIPO},{T_TIPOS},5,FALSE)",
            f"={F_COMPLEX}",
            f"={P}$B$7",
        ),
        (
            "Integrações / API",
            f"=VLOOKUP({IN_TIPO},{T_TIPOS},6,FALSE)+{IN_INTEGR}*{P}$B$22",
            f"={F_COMPLEX}",
            f"={P}$B$8",
        ),
        (
            "QA / Testes",
            f"=VLOOKUP({IN_TIPO},{T_TIPOS},7,FALSE)",
            f"={F_COMPLEX}*{F_IDIOMAS}",
            f"={P}$B$9",
        ),
        (
            "Produção de conteúdos",
            f"=VLOOKUP({IN_CONT},{T_CONT},2,FALSE)",
            f"={F_IDIOMAS}",
            f"={P}$B$11",
        ),
        ("SEO", f"=VLOOKUP({IN_SEO},{T_SEO},2,FALSE)", f"={F_IDIOMAS}", f"={P}$B$12"),
    ]

    for i, (nome, horas, fator, tarifa) in enumerate(fases):
        r = 5 + i
        ws.cell(row=r, column=1, value=nome).font = f_label
        ws.cell(row=r, column=2, value=horas).number_format = HORAS
        ws.cell(row=r, column=3, value=fator).number_format = FATOR
        ws.cell(row=r, column=4, value=f"=B{r}*C{r}").number_format = HORAS
        ws.cell(row=r, column=5, value=tarifa).number_format = EUR
        ws.cell(row=r, column=6, value=f"=D{r}*E{r}").number_format = EUR2
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = box
            if i % 2:
                ws.cell(row=r, column=col).fill = fill_zebra

    linhas_totais = [
        (13, "Subtotal técnico", "=SUM(D5:D12)", None, "=SUM(F5:F12)"),
        (14, "Gestão de projeto", f"=D13*{P}$B$15", f"={P}$B$10", "=D14*E14"),
        (15, "Buffer de risco", f"=(D13+D14)*{P}$B$16", None, f"=(F13+F14)*{P}$B$16"),
        (16, "TOTAL", "=D13+D14+D15", None, "=F13+F14+F15"),
    ]
    for r, nome, horas, tarifa, valor in linhas_totais:
        ws.cell(row=r, column=1, value=nome).font = f_bold
        ws.cell(row=r, column=4, value=horas).number_format = HORAS
        if tarifa:
            ws.cell(row=r, column=5, value=tarifa).number_format = EUR
        ws.cell(row=r, column=6, value=valor).number_format = EUR2
        for col in range(1, 7):
            cell = ws.cell(row=r, column=col)
            cell.border = box
            cell.fill = fill_sub
            cell.font = f_bold

    ajustes = [
        (18, "Fator de urgência", f"=VLOOKUP({IN_URG},{T_URG},2,FALSE)"),
        (19, "Fator de perfil de cliente", f"=VLOOKUP({IN_PERFIL},{T_CLIENTE},2,FALSE)"),
    ]
    for r, nome, formula in ajustes:
        ws.cell(row=r, column=1, value=nome).font = f_label
        c = ws.cell(row=r, column=4, value=formula)
        c.number_format = FATOR
        c.border = box

    ws.cell(row=20, column=1, value="PREÇO BASE (antes de escalões)").font = f_bold
    c = ws.cell(row=20, column=6, value="=F16*D18*D19")
    c.number_format = EUR2
    c.font = f_big
    c.fill = fill_out
    c.border = box

    ws.column_dimensions["A"].width = 30
    for col in "BCDEF":
        ws.column_dimensions[col].width = 17


def build_calculadora(ws):
    ws.sheet_view.showGridLines = False
    ws["B1"] = "CALCULADORA DE PREÇOS esDEV"
    ws["B1"].font = f_title
    ws["B2"] = "Preenche apenas as células amarelas. Tudo o resto é calculado."
    ws["B2"].font = f_hint

    secao(ws, 4, "1. CARACTERÍSTICAS DO PROJETO", ncols=4, col=2)

    inputs = [
        (5, "Cliente / Projeto", DEFAULTS["cliente"], "@", None),
        (6, "Tipo de projeto", DEFAULTS["tipo"], "@", f"={P}$A$27:$A$33"),
        (7, "Nível de complexidade", DEFAULTS["complexidade"], "@", f"={P}$A$36:$A$39"),
        (8, "Nº de páginas / ecrãs", DEFAULTS["paginas"], "0", None),
        (9, "Nível de design", DEFAULTS["design"], "@", f"={P}$A$42:$A$44"),
        (10, "Nº de idiomas", DEFAULTS["idiomas"], "0", None),
        (11, "Nº de integrações externas", DEFAULTS["integracoes"], "0", None),
        (12, "Conteúdos (textos / imagens)", DEFAULTS["conteudos"], "@", f"={P}$A$57:$A$58"),
        (13, "SEO", DEFAULTS["seo"], "@", f"={P}$A$61:$A$63"),
        (14, "Prazo / urgência", DEFAULTS["urgencia"], "@", f"={P}$A$47:$A$49"),
        (15, "Perfil de cliente", DEFAULTS["perfil"], "@", f"={P}$A$52:$A$54"),
        (16, "Desconto comercial", DEFAULTS["desconto"], PCT, None),
        (17, "Custos externos a repassar", DEFAULTS["custos"], EUR, None),
    ]

    dicas = {
        6: "Define as horas base de cada fase.",
        8: "Acima das páginas incluídas, cada página soma horas.",
        10: "Cada idioma extra acresce % em frontend, QA, conteúdos e SEO.",
        11: "Pagamentos, CRM, ERP, APIs, newsletter…",
        16: "Usar com critério: sai direto da margem.",
        17: "Licenças, fotografia, domínio, 1º ano de alojamento.",
    }

    for row, nome, valor, fmt, lista in inputs:
        rotulo(ws, row, nome, dicas.get(row))
        c = ws.cell(row=row, column=3, value=valor)
        c.fill = fill_input
        c.font = f_bold
        c.number_format = fmt
        c.border = box
        c.alignment = Alignment(horizontal="left")
        if lista:
            dv = DataValidation(type="list", formula1=lista, allow_blank=False)
            dv.error = "Escolhe uma opção da lista (definida na folha Parametros)."
            ws.add_data_validation(dv)
            dv.add(c)

    secao(ws, 19, "2. ESFORÇO E PRAZO", ncols=4, col=2)
    esforco = [
        (20, "Horas estimadas (com gestão e buffer)", "=Detalhe!D16", HORAS),
        (21, "Prazo de produção estimado", f"=ROUNDUP(Detalhe!D16/{P}$B$17,0)", '0 " semanas"'),
        (22, "Preço base do modelo", "=Detalhe!F20", EUR2),
    ]
    for row, nome, formula, fmt in esforco:
        rotulo(ws, row, nome)
        c = ws.cell(row=row, column=3, value=formula)
        c.number_format = fmt
        c.font = f_bold
        c.border = box

    secao(ws, 24, "3. ESCALÕES DE PREÇO", ncols=4, col=2)
    for j, titulo in enumerate(["Escalão", "Sem IVA", "Com IVA", "Quando usar"]):
        c = ws.cell(row=25, column=2 + j, value=titulo)
        c.font = f_sub
        c.fill = fill_sub
        c.border = box

    preco_min_abs = f"VLOOKUP($C$6,{T_TIPOS},9,FALSE)"

    def formula_escalao(fator: str) -> str:
        bruto = f"ROUND($C$22*{fator}/{P}$B$23,0)*{P}$B$23"
        return f"=MAX({preco_min_abs},{bruto})*(1-$C$16)+$C$17"

    escaloes = [
        (
            26,
            "Preço Mínimo",
            formula_escalao(f"{P}$B$18"),
            "Piso de negociação. Abaixo disto, recusar o projeto.",
        ),
        (
            27,
            "Preço Recomendado",
            formula_escalao("1"),
            "Valor a apresentar por defeito na proposta.",
        ),
        (
            28,
            "Preço Premium",
            formula_escalao(f"{P}$B$19"),
            "Prioridade na fila, revisões extra e acompanhamento próximo.",
        ),
    ]
    for row, nome, formula, nota in escaloes:
        c = ws.cell(row=row, column=2, value=nome)
        c.font = f_bold
        c.border = box
        c = ws.cell(row=row, column=3, value=formula)
        c.number_format = EUR
        c.border = box
        c.fill = fill_out
        c.font = f_big if row == 27 else f_bold
        c = ws.cell(row=row, column=4, value=f"=C{row}*(1+{P}$B$24)")
        c.number_format = EUR
        c.border = box
        c = ws.cell(row=row, column=5, value=nota)
        c.font = f_hint
        c.alignment = Alignment(wrap_text=True, vertical="center")

    rotulo(ws, 29, "Valor / hora efetivo (Recomendado)")
    c = ws.cell(row=29, column=3, value="=IFERROR(($C$27-$C$17)/$C$20,0)")
    c.number_format = EUR2
    c.font = f_bold
    c.border = box
    ws.cell(row=29, column=5, value="Comparar com a tarifa alvo. Se descer muito, o desconto é excessivo.").font = f_hint

    secao(ws, 31, "4. MANUTENÇÃO MENSAL", ncols=4, col=2)
    for i, (nome, _, _, inclui) in enumerate(MANUTENCAO):
        row = 32 + i
        pr = 66 + i
        c = ws.cell(row=row, column=2, value=f"Plano {nome}")
        c.font = f_bold
        c.border = box
        c = ws.cell(row=row, column=3, value=f"=MAX({P}$C${pr},ROUND($C$27*{P}$B${pr},0))")
        c.number_format = '#,##0 "€ / mês"'
        c.font = f_bold
        c.fill = fill_out
        c.border = box
        c = ws.cell(row=row, column=4, value=f"=C{row}*12")
        c.number_format = '#,##0 "€ / ano"'
        c.border = box
        c = ws.cell(row=row, column=5, value=inclui)
        c.font = f_hint
        c.alignment = Alignment(wrap_text=True, vertical="center")

    secao(ws, 36, "5. PLANO DE PAGAMENTO (sobre o Preço Recomendado)", ncols=4, col=2)
    pagamentos = [
        (37, "Sinal — arranque", 0.40),
        (38, "Aprovação do design / staging", 0.30),
        (39, "Entrega e publicação", 0.30),
    ]
    for row, nome, pct in pagamentos:
        rotulo(ws, row, nome)
        c = ws.cell(row=row, column=3, value=pct)
        c.number_format = PCT
        c.fill = fill_input
        c.border = box
        c = ws.cell(row=row, column=4, value=f"=$C$27*C{row}")
        c.number_format = EUR
        c.font = f_bold
        c.border = box

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 58
    for r in (26, 27, 28, 32, 33, 34):
        ws.row_dimensions[r].height = 30
    ws.freeze_panes = "A4"


def build_proposta(ws):
    ws.sheet_view.showGridLines = False
    ws["B1"] = "RESUMO PARA PROPOSTA"
    ws["B1"].font = f_title
    ws["B2"] = "Pronto a copiar para o e-mail ou documento de proposta."
    ws["B2"].font = f_hint

    linhas = [
        ("Projeto", "=Calculadora!$C$5"),
        ("Âmbito", "=Calculadora!$C$6&\" · \"&Calculadora!$C$7"),
        (
            "Dimensão",
            '=Calculadora!$C$8&" páginas/ecrãs · "&Calculadora!$C$10&" idioma(s) · "&Calculadora!$C$11&" integração(ões)"',
        ),
        ("Design", "=Calculadora!$C$9"),
        ("Conteúdos", "=Calculadora!$C$12"),
        ("SEO", "=Calculadora!$C$13"),
        ("Esforço estimado", '=TEXT(Calculadora!$C$20,"#,##0")&" horas"'),
        ("Prazo de produção", '=Calculadora!$C$21&" semanas após aprovação"'),
    ]
    for i, (nome, formula) in enumerate(linhas):
        r = 4 + i
        ws.cell(row=r, column=2, value=nome).font = f_bold
        c = ws.cell(row=r, column=3, value=formula)
        c.font = f_label
        c.border = box

    secao(ws, 13, "INVESTIMENTO", ncols=3, col=2)
    ws.cell(row=14, column=2, value="Valor do projeto (s/ IVA)").font = f_bold
    c = ws.cell(row=14, column=3, value="=Calculadora!$C$27")
    c.number_format = EUR
    c.font = f_big
    c.fill = fill_out
    c.border = box

    ws.cell(row=15, column=2, value="Valor com IVA à taxa legal").font = f_label
    c = ws.cell(row=15, column=3, value="=Calculadora!$D$27")
    c.number_format = EUR
    c.border = box

    ws.cell(row=16, column=2, value="Manutenção mensal recomendada (Pro)").font = f_label
    c = ws.cell(row=16, column=3, value="=Calculadora!$C$33")
    c.number_format = '#,##0 "€ / mês"'
    c.border = box

    ws.cell(row=18, column=2, value="Condições de pagamento").font = f_bold
    ws.cell(
        row=19,
        column=2,
        value='=TEXT(Calculadora!$D$37,"#,##0 €")&" no arranque · "&TEXT(Calculadora!$D$38,"#,##0 €")&" na aprovação do design · "&TEXT(Calculadora!$D$39,"#,##0 €")&" na entrega"',
    ).font = f_label

    ws.cell(row=21, column=2, value="Notas").font = f_bold
    notas = [
        "Proposta válida por 30 dias.",
        "Inclui duas rondas de revisão por fase; revisões adicionais são orçamentadas à hora.",
        "Alterações de âmbito após aprovação seguem novo orçamento.",
        "Licenças, alojamento e domínio são custos de terceiros, faturados a preço de custo.",
    ]
    for i, nota in enumerate(notas):
        ws.cell(row=22 + i, column=2, value="• " + nota).font = f_label

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 62


def main(destino: str) -> None:
    wb = Workbook()
    ws_calc = wb.active
    ws_calc.title = "Calculadora"
    ws_det = wb.create_sheet("Detalhe")
    ws_par = wb.create_sheet("Parametros")
    ws_prop = wb.create_sheet("Proposta")

    build_parametros(ws_par)
    build_detalhe(ws_det)
    build_calculadora(ws_calc)
    build_proposta(ws_prop)

    for ws in wb.worksheets:
        ws.sheet_properties.tabColor = AZUL
    wb.active = 0
    wb.save(destino)
    print(f"Gerado: {destino}")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "Calculadora_Precos_esDEV.xlsx")
